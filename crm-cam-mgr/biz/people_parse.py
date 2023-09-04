import os
from datetime import datetime
from mtkext.guid import fast_guid
import aiofiles
from sanic.log import logger
from common.models.cam import PeopleFile, UserWhiteList
from mtkext.db import FlexModel



async def file_upload_api(app, params):
    db = app.mgr
    files = params.get("files")
    apply_type = params.get("type")
    instance_id = params.get("instance_id")
    file_obj = await upload_file(app, files)
    if not file_obj:
        return 0, dict(msg="上传保存数据异常")
    file_obj.update({"from_type": 1, "apply_type": apply_type, "instance_id":instance_id})
    pf_obj = await db.create(PeopleFile, **file_obj)
    result = {"people_ids": [pf_obj.auto_id]}
    file_path = file_obj.get("file_path")
    member_nos = await read_file(file_path, "member_no")
    logger.info(f"member_nos: {member_nos}")
    if member_nos:
        rows, plus = divmod(len(member_nos), 10000)
        iter_max = rows + 1 if plus else rows
        logger.info(f"iter_max: {iter_max}")
        cls = FlexModel.get(UserWhiteList, pf_obj.auto_id)
        for i in range(0, iter_max):
            start = i * 10000
            end = start + 10000
            items = member_nos[start: end]
            _items = [{
                "member_no": x,
                "instance_id": instance_id,
            } for x in items]
            got = await db.execute(cls.insert_many(items).on_conflict_ignore())
            logger.info(f"insert many got: {got}")

        if plus:
            logger.info(f"plus: {-plus}")
            items = member_nos[-plus:]
            logger.info(f"items:{items}")
            _items = [{
                "member_no": x,
                "instance_id": instance_id,
            } for x in items]
            got = await db.execute(cls.insert_many(items).on_conflict_ignore())
            logger.info(f"insert many got: {got}")
    else:
        res = []
    return 1, dict(data=result)




async def get_people_info(app, people_ids):
    # TODO 数据量过大时优化
    db = app.mgr
    event_db = app.event_db
    people_files = await db.execute(PeopleFile.select().where(PeopleFile.id.in_(people_ids)).dicts())
    result = set()
    for item in people_files:
        from_type = item.get("from_type")
        apply_type = item.get("apply_type")
        file_path = item.get("file_path")
        if from_type == 1:
            res = await read_file(file_path, "Sheet1")
            if res:
                result.update(set(res))
        else:
            if apply_type == 1:
                res = await read_file(file_path, "mobile")
            else:
                unionds = await read_file(file_path, "unionid")
                if unionds:
                    rows, plus = divmod(len(unionds), 1000)
                    res = []
                    iter_max = rows + 1 if plus else rows
                    for i in range(0, iter_max):
                        start = i * 1000
                        end = start + 1000
                        items = unionds[start: end]
                        openid_obj_list = await event_db.execute(
                            UserSubscribeEvent.select(UserSubscribeEvent.openid).where(
                                UserSubscribeEvent.unionid.in_(items)).dicts())
                        res.extend([item.get("openid") for item in openid_obj_list])
                else:
                    res = []
            if res:
                result.update(set(res))
    return list(result)


async def read_file(file_path, sheet_name):
    res = set()
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    # sheet1_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        item = row[0]
        if item and item not in ["phone", "openid", "phones", "openids", "member_no"]:
            res.add(str(item))
    return list(res)


async def upload_file(app, files):
    base_path, file_folder = await check_dir(app)
    file_obj = {}
    try:
        for _, form_files in files.items():
            for file in form_files:
                ext = file.name.split(".")[-1]
                file_name = f"{fast_guid()}.{ext}"
                file_path = f"{base_path}/{file_folder}/{file_name}"
                file_body = file.body
                async with aiofiles.open(file_path, mode='wb') as fp:
                    await fp.write(file_body)
                file_obj = dict(file_name=file_name, file_path=file_path)
                break
            break
    except Exception as e:
        logger.error(f"upload_file error")
        logger.exception(e)
    finally:
        return file_obj


async def download_file(app, url):
    http = app.http
    base_path, file_folder = await check_dir(app)
    file_name = f"{fast_guid}.xlsx"
    file_path = f"{base_path}/{file_folder}/{file_name}"
    file_obj = {}
    try:
        logger.info(f"download_file request => {url}")
        res = await http.get(url, parse_with_json=False)
        logger.info(f"download_file response => {url, res}")
        if res:
            async with aiofiles.open(file_path, 'wb') as fd:
                await fd.write(res)
            file_obj = dict(file_name=file_name, file_path=file_path)
    except Exception as e:
        logger.error(f"download file error: {url}")
        logger.exception(e)
    finally:
        return file_obj


async def check_dir(app):
    conf = app.conf
    base_path = conf.PEOPLE_FILE_DIR
    file_folder = f"{datetime.now().strftime('%Y%m%d')}"
    if not os.path.exists(os.path.join(base_path, file_folder)):
        os.makedirs(os.path.join(base_path, file_folder))
    return base_path, file_folder
